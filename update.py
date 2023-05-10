import openpyxl

def update():
    print("####################   UPDATE PASSENGER DETAILS  ###########################")
    a=int(input("enter TICKET NUMBER to update details :"))
    mo=openpyxl.load_workbook(r"C:\\Users\\dell\\PycharmProjects\\TICKETBOOKINGG\\TRICKET_BOOK.xlsx")
    list=mo.get_sheet_names()
    a1=mo[list[1]]
    c=0

    for i in a1:
        for j in i:
            pass
        c+=1

    if a<c:
        l = ['ticket number', 'passenger name', 'boarding place', 'destination', 'travel date', 'travel time', 'fare',
             'booking status']
        for i in range(len(l)):
            if i==0:
                a1.cell(a+1,i+1).value=a
            if i==len(l)-1:
                a1.cell(a+1,i+1).value=1
            if 1<= i <= len(l)-2:
                a1.cell(a+1,i+1).value=input(l[i])
        mo.save(r"C:\\Users\\dell\\PycharmProjects\\TICKETBOOKINGG\\TRICKET_BOOK.xlsx")
        print("#############   PASSENGER DETAILES UPDATED SUCCESFULLY   #############")
        print()
    else:
        print("#############   ENTERED TICKET NUMBER NOT FOUND ##############")
        print()