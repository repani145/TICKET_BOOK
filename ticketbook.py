import openpyxl

def initital():
    mo=openpyxl.Workbook()
    sh1=mo.create_sheet()

    l=['ticket number','passenger name','boarding place','destination','travel date','travel time','fare','booking status']
    n=len(l)

    for i in range(1,2):
        for j in range(n):
            sh1.cell(i,j+1).value=l[j]
    mo.save(r"C:\\Users\\dell\PycharmProjects\\TICKETBOOKINGG\\TRICKET_BOOK.xlsx")

def ticket_booking():
    try:
        print("#################### TICKET BOOKING ############################")
        b = ['ticket number', 'passenger name', 'boarding place', 'destination', 'travel date', 'travel time', 'fare',
             'booking status']
        m=len(b)
        ob=openpyxl.load_workbook(r"C:\\Users\\dell\\PycharmProjects\\TICKETBOOKINGG\\TRICKET_BOOK.xlsx")
        list = ob.get_sheet_names()
        a1=ob[list[1]]
        c=0
        for i in a1:
            for j in i:
                pass
            c+=1
        print(c)

        for i in range(m):
            if i==0:
                a1.cell(c+1,i+1).value=c
            if 1 <= i <= m-2 :
                a1.cell(c + 1, i + 1).value = input(b[i])
            if i==m-1:
                a1.cell(c + 1, i + 1).value = 1

        ob.save(r"C:\\Users\dell\\PycharmProjects\\TICKETBOOKINGG\\TRICKET_BOOK.xlsx")


    except Exception as ok:
        print("#########   ticket booking not completed     ############")
    else:

        print('##############    TICKET BOOKED SUCCESSFULLY     ###############')

    print("########################################################################################################")








